from openpyxl.workbook import Workbook
from project_dataclasses import Person, Vehicle, HouseholdItem


def write_people_xlsx(people: list[Person], workbook: Workbook, sheet_name: str = "People") -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["ID", "Name", "Age", "Male"])
    for person in people:
        sheet.append([person.id, person.name, person.age, person.male])


def read_people_xlsx(workbook: Workbook, sheet_name: str = "People") -> list[Person]:
    sheet = workbook[sheet_name]
    return [Person(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value,
                   sheet.cell(row=row, column=3).value, sheet.cell(row=row, column=4).value)
            for row in range(2, sheet.max_row + 1)]


def write_vehicles_xlsx(vehicles: list[Vehicle], workbook: Workbook, sheet_name: str = "Vehicles") -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["ID", "Type", "Model", "Owner ID"])
    for vehicle in vehicles:
        sheet.append([vehicle.id, vehicle.type, vehicle.model, vehicle.owner_id])


def read_vehicles_xlsx(workbook: Workbook, sheet_name: str = "Vehicles") -> list[Vehicle]:
    sheet = workbook[sheet_name]
    return [Vehicle(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value, sheet.cell(row=row, column=4).value)
            for row in range(2, sheet.max_row + 1)]


def write_household_items_xlsx(household_items: list[HouseholdItem], workbook: Workbook, sheet_name: str = "HouseholdItems") -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["ID", "Name", "Category", "Owner ID"])
    for item in household_items:
        sheet.append([item.id, item.name, item.category, item.owner_id])


def read_household_items_xlsx(workbook: Workbook, sheet_name: str = "HouseholdItems") -> list[HouseholdItem]:
    sheet = workbook[sheet_name]
    return [HouseholdItem(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value,
                          sheet.cell(row=row, column=3).value, sheet.cell(row=row, column=4).value)
            for row in range(2, sheet.max_row + 1)]
