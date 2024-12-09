import os
from openpyxl import Workbook, load_workbook
from generator import generate_people, generate_vehicles, generate_household_items
from XLSX import (
    write_people_xlsx, write_vehicles_xlsx, write_household_items_xlsx,
    read_people_xlsx, read_vehicles_xlsx, read_household_items_xlsx
)
from CSV import (
    write_people, write_vehicles, write_household_items,
    read_people, read_vehicles, read_household_items
)
from JSON import (
    write_people_json, write_vehicles_json, write_household_items_json,
    read_people_json, read_vehicles_json, read_household_items_json
)


def main():
    num_entries = 50
    people = generate_people(num_entries)
    vehicles = generate_vehicles(num_entries, people)
    household_items = generate_household_items(num_entries, people)

    path = "data"
    os.makedirs(path, exist_ok=True)

    write_people(people, path)
    write_vehicles(vehicles, path)
    write_household_items(household_items, path)

    write_people_json(people, path)
    write_vehicles_json(vehicles, path)
    write_household_items_json(household_items, path)

    workbook = Workbook()
    write_people_xlsx(people, workbook)
    write_vehicles_xlsx(vehicles, workbook)
    write_household_items_xlsx(household_items, workbook)
    workbook.save(os.path.join(path, "data.xlsx"))

    people_from_csv = read_people(path)
    vehicles_from_csv = read_vehicles(path)
    household_items_from_csv = read_household_items(path)

    print("CSV Sample:", people_from_csv[:3], vehicles_from_csv[:3], household_items_from_csv[:3])

    workbook_from_xlsx = load_workbook(os.path.join(path, "data.xlsx"))
    people_from_xlsx = read_people_xlsx(workbook_from_xlsx)
    vehicles_from_xlsx = read_vehicles_xlsx(workbook_from_xlsx)
    household_items_from_xlsx = read_household_items_xlsx(workbook_from_xlsx)

    print("XLSX Sample:", people_from_xlsx[:3], vehicles_from_xlsx[:3], household_items_from_xlsx[:3])

    people_from_json = read_people_json(path)
    vehicles_from_json = read_vehicles_json(path)
    household_items_from_json = read_household_items_json(path)

    print("\nJSON Sample:")
    print("People from JSON:", people_from_json[:3])
    print("Vehicles from JSON:", vehicles_from_json[:3])
    print("Household Items from JSON:", household_items_from_json[:3])


if __name__ == "__main__":
    main()
